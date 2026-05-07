#!/usr/bin/env python3
"""
parse_excel.py — QA Agent Skill v2
Parses manual test case Excel sheets into structured JSON.

Designed for the Common Ground test case format:
  Col A: Sr. No.
  Col B: Test Case ID
  Col C: Persona
  Col D: Test Case Scenario
  Col E: Preconditions
  Col F: Test Data (ignored — data generated dynamically)
  Col G: Test Steps Sr. No
  Col H: Test Steps
  Col I: Expected Result
  Col J+: Ignored

Handles merged cells: TC metadata appears only on the first row of each
test case. Subsequent rows carry only step_no, step_text, expected_result.

Usage:
  python parse_excel.py --input "path/to/tests.xlsx" --output parsed.json --sheet 0
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


# ─── Step classification ─────────────────────────────────────────────────────

VALIDATION_KEYWORDS = [
    'validate', 'verify', 'check', 'confirm', 'ensure', 'assert',
    'should be', 'should have', 'should display', 'should show',
    'must be', 'must have', 'is visible', 'is present', 'is displayed',
    'is enabled', 'is disabled', 'is mandatory', 'is not',
]

ACTION_KEYWORDS = [
    'click', 'clicks', 'press', 'tap', 'select', 'selects',
    'enter', 'enters', 'type', 'types', 'fill', 'fills',
    'navigate', 'navigates', 'open', 'opens', 'go to',
    'scroll', 'scrolls', 'drag', 'drop', 'upload', 'uploads',
    'hover', 'hovers', 'submit', 'submits', 'save', 'saves',
    'clear', 'clears', 'delete', 'deletes', 'remove',
    'toggle', 'switch', 'expand', 'collapse',
    'user is present',  # precondition step — treated as action (navigate)
]


def classify_step(step_text: str) -> str:
    """Classify a step as 'action' or 'validation' based on keywords."""
    lower = step_text.lower().strip()

    # Check validation first — "Validate that..." is clearly an assertion
    for kw in VALIDATION_KEYWORDS:
        if kw in lower:
            return 'validation'

    # Then check action
    for kw in ACTION_KEYWORDS:
        if kw in lower:
            return 'action'

    # Default: if step starts with "User" it's likely an action
    if lower.startswith('user'):
        return 'action'

    # If step starts with a noun/the, it might be a state observation
    return 'validation'


# ─── UI element extraction from step text ────────────────────────────────────

def extract_ui_element(step_text: str) -> dict:
    """Extract the UI element reference from a step's natural language."""
    result = {
        'quoted_text': None,     # text in quotes: "Add Indicator"
        'element_type': None,    # button, dropdown, textbox, link, etc.
        'location_hint': None,   # "top right corner", "left pane", etc.
    }

    # Extract quoted text — these are usually button labels or field names
    quotes = re.findall(r'"([^"]+)"', step_text)
    if quotes:
        result['quoted_text'] = quotes[0]

    # Detect element type
    lower = step_text.lower()
    if any(w in lower for w in ['button', 'btn']):
        result['element_type'] = 'button'
    elif any(w in lower for w in ['dropdown', 'drop-down', 'drop down', 'select']):
        result['element_type'] = 'dropdown'
    elif any(w in lower for w in ['textbox', 'text box', 'text field', 'input field', 'input box']):
        result['element_type'] = 'textbox'
    elif any(w in lower for w in ['checkbox', 'check box']):
        result['element_type'] = 'checkbox'
    elif any(w in lower for w in ['radio', 'radio button']):
        result['element_type'] = 'radio'
    elif any(w in lower for w in ['link', 'hyperlink']):
        result['element_type'] = 'link'
    elif any(w in lower for w in ['tab', 'pane', 'panel']):
        result['element_type'] = 'tab'
    elif any(w in lower for w in ['icon']):
        result['element_type'] = 'icon'
    elif any(w in lower for w in ['popup', 'pop-up', 'modal', 'dialog']):
        result['element_type'] = 'modal'
    elif 'enter' in lower or 'type' in lower or 'fill' in lower:
        result['element_type'] = 'textbox'

    # Location hints
    for loc in ['top right', 'top left', 'bottom right', 'bottom left',
                'left nav', 'left pane', 'right side', 'top bar',
                'navigational pane', 'header', 'footer', 'sidebar']:
        if loc in lower:
            result['location_hint'] = loc
            break

    return result


# ─── Main parser ─────────────────────────────────────────────────────────────

def parse_workbook(filepath: str, sheet_index: int = 0) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = wb.sheetnames[sheet_index]
    ws = wb[sheet_name]
    print(f"  Reading sheet: '{sheet_name}'", file=sys.stderr)

    result = {
        'source_file': Path(filepath).name,
        'sheet': sheet_name,
        'personas': [],
        'modules': [],
        'test_cases': [],
    }

    personas_seen = set()
    modules_seen  = set()
    tc_ids_seen   = {}   # tc_id_str -> row_number of first occurrence

    # State for merged cell handling
    current_tc = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Column mapping (0-indexed)
        sr_no          = row[0] if len(row) > 0 else None
        tc_id          = row[1] if len(row) > 1 else None
        persona        = row[2] if len(row) > 2 else None
        scenario       = row[3] if len(row) > 3 else None
        preconditions  = row[4] if len(row) > 4 else None
        _test_data     = row[5] if len(row) > 5 else None   # ignored
        step_no        = row[6] if len(row) > 6 else None
        step_text      = row[7] if len(row) > 7 else None
        expected       = row[8] if len(row) > 8 else None

        # ── New test case row (has tc_id) ─────────────────────────────────
        if tc_id and str(tc_id).strip():
            tc_id_str    = str(tc_id).strip()
            persona_str  = str(persona or '').strip()
            scenario_str = str(scenario or '').strip()
            precon_str   = str(preconditions or '').strip()

            # WARNING: duplicate TC ID -- later entry silently replaces earlier one
            if tc_id_str in tc_ids_seen:
                print(
                    f"WARNING: Duplicate TC ID '{tc_id_str}' at row {row_idx} "
                    f"(first seen at row {tc_ids_seen[tc_id_str]}). "
                    f"The earlier test case will be replaced by this one.",
                    file=sys.stderr,
                )
            tc_ids_seen[tc_id_str] = row_idx

            # Derive module from TC ID prefix: "IM-E2E-001" -> "Indicator Management"
            module = derive_module(tc_id_str, scenario_str)

            # WARNING: module could not be derived from prefix -- fell back to 'General'
            if module == 'General':
                print(
                    f"WARNING: Could not derive module for TC ID '{tc_id_str}' "
                    f"(prefix '{tc_id_str.split('-')[0]}' not in module_map and "
                    f"no keyword matched in scenario). Assigned module: 'General'. "
                    f"Update derive_module() or check the TC ID prefix.",
                    file=sys.stderr,
                )

            current_tc = {
                'id':            tc_id_str,
                'persona':       persona_str,
                'module':        module,
                'scenario':      scenario_str,
                'preconditions': precon_str,
                'steps':         [],
            }
            # Replace any earlier duplicate entry in test_cases list
            result['test_cases'] = [tc for tc in result['test_cases'] if tc['id'] != tc_id_str]
            result['test_cases'].append(current_tc)

            if persona_str and persona_str not in personas_seen:
                personas_seen.add(persona_str)
                result['personas'].append(persona_str)

            if module and module not in modules_seen:
                modules_seen.add(module)
                result['modules'].append(module)

        # ── Step row ──────────────────────────────────────────────────────
        # WARNING: row has a step number but empty step text -- possible data entry gap
        if current_tc and step_no is not None and not (step_text and str(step_text).strip()):
            print(
                f"WARNING: Row {row_idx} has step number '{step_no}' but empty step text. "
                f"Skipping row. Check TC '{current_tc['id']}' in the Excel for missing data.",
                file=sys.stderr,
            )

        if current_tc and step_text and str(step_text).strip():
            step_str     = str(step_text).strip()
            expected_str = str(expected or '').strip()
            step_type    = classify_step(step_str)
            ui_element   = extract_ui_element(step_str)

            current_tc['steps'].append({
                'step_no':     float(step_no) if step_no else len(current_tc['steps']) + 1,
                'step_text':   step_str,
                'expected':    expected_str,
                'type':        step_type,      # 'action' or 'validation'
                'ui_element':  ui_element,
            })

    wb.close()

    # Summary
    total_steps = sum(len(tc['steps']) for tc in result['test_cases'])
    action_steps = sum(
        1 for tc in result['test_cases']
        for s in tc['steps'] if s['type'] == 'action'
    )
    validation_steps = total_steps - action_steps

    result['summary'] = {
        'total_test_cases': len(result['test_cases']),
        'total_steps':      total_steps,
        'action_steps':     action_steps,
        'validation_steps': validation_steps,
        'personas':         result['personas'],
        'modules':          result['modules'],
    }

    print(f"  Parsed: {len(result['test_cases'])} test cases, {total_steps} steps "
          f"({action_steps} actions, {validation_steps} validations)", file=sys.stderr)
    print(f"  Personas: {result['personas']}", file=sys.stderr)
    print(f"  Modules: {result['modules']}", file=sys.stderr)

    return result


def derive_module(tc_id: str, scenario: str) -> str:
    """Derive module name from TC ID prefix or scenario text."""
    prefix = tc_id.split('-')[0].upper() if '-' in tc_id else ''

    module_map = {
        'IM':  'Indicator Management',
        'RM':  'Result Management',
        'RES': 'Result Management',
        'AUTH': 'Authentication',
        'USR': 'User Management',
        'RPT': 'Reports',
        'DAS': 'Dashboard',
    }

    if prefix in module_map:
        return module_map[prefix]

    # Fallback: extract from scenario text
    lower = scenario.lower()
    if 'indicator' in lower: return 'Indicator Management'
    if 'result' in lower:    return 'Result Management'
    if 'login' in lower:     return 'Authentication'
    if 'report' in lower:    return 'Reports'

    return 'General'


def main():
    parser = argparse.ArgumentParser(description="Parse manual test case Excel")
    parser.add_argument("--input",  required=True, help="Path to Excel file")
    parser.add_argument("--output", required=True, help="Path to output JSON")
    parser.add_argument("--sheet",  type=int, default=0, help="Sheet index (0-based)")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"Error: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {args.input}", file=sys.stderr)
    data = parse_workbook(args.input, args.sheet)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Output: {args.output}", file=sys.stderr)

    # Print summary to stdout for Claude to read
    print(json.dumps(data['summary'], indent=2))


if __name__ == '__main__':
    main()
